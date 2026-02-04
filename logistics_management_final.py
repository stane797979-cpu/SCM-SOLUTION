# -*- coding: utf-8 -*-
import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from datetime import datetime
import openpyxl

st.set_page_config(page_title="물류비 관리", page_icon="💰", layout="wide")

st.markdown("""
<style>
.main-header {font-size: 2.5rem; font-weight: bold; color: #1F4E78; text-align: center; padding: 1rem 0;}
.alert-danger {background-color: #FFC7CE; padding: 1rem; border-radius: 0.5rem; border-left: 4px solid #C00000; margin: 1rem 0;}
.alert-warning {background-color: #FFEB9C; padding: 1rem; border-radius: 0.5rem; border-left: 4px solid #FFA500; margin: 1rem 0;}
.alert-success {background-color: #C6EFCE; padding: 1rem; border-radius: 0.5rem; border-left: 4px solid #006100; margin: 1rem 0;}
</style>
""", unsafe_allow_html=True)

@st.cache_data
def load_logistics_data(file_path):
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)

        ws = wb['회사정보']
        company = {
            '회사명': ws.cell(2, 2).value,
            '업종': ws.cell(3, 2).value,
            '연매출': ws.cell(4, 2).value,
            '목표물류비율': ws.cell(5, 2).value
        }

        ws = wb['월별물류비']
        monthly = []
        for row in range(3, ws.max_row + 1):
            ym = ws.cell(row, 1).value
            if ym:
                monthly.append({
                    '년월': ym,
                    '국제운송비': ws.cell(row, 2).value or 0,
                    '창고비': ws.cell(row, 3).value or 0,
                    '국내배송비': ws.cell(row, 4).value or 0,
                    '포장비': ws.cell(row, 5).value or 0,
                    '통관비': ws.cell(row, 6).value or 0,
                    '기타': ws.cell(row, 7).value or 0
                })
        df_monthly = pd.DataFrame(monthly)
        df_monthly['총물류비'] = df_monthly[['국제운송비', '창고비', '국내배송비', '포장비', '통관비', '기타']].sum(axis=1)
        df_monthly['물류비율'] = (df_monthly['총물류비'] / company['연매출'] * 100)

        ws = wb['협력사']
        partners = []
        for row in range(3, ws.max_row + 1):
            name = ws.cell(row, 1).value
            if name:
                partners.append({
                    '협력사명': name,
                    '구분': ws.cell(row, 2).value,
                    '월평균비용': ws.cell(row, 3).value or 0,
                    '계약시작일': ws.cell(row, 4).value,
                    '계약종료일': ws.cell(row, 5).value,
                    '담당자': ws.cell(row, 6).value,
                    '연락처': ws.cell(row, 7).value
                })
        df_partners = pd.DataFrame(partners)

        ws = wb['개선과제']
        tasks = []
        for row in range(3, ws.max_row + 1):
            task = ws.cell(row, 1).value
            if task:
                tasks.append({
                    '과제명': task,
                    '카테고리': ws.cell(row, 2).value,
                    '상태': ws.cell(row, 3).value,
                    '목표절감액': ws.cell(row, 4).value or 0,
                    '실제절감액': ws.cell(row, 5).value or 0,
                    '담당자': ws.cell(row, 6).value,
                    '시작일': ws.cell(row, 7).value,
                    '완료예정일': ws.cell(row, 8).value
                })
        df_tasks = pd.DataFrame(tasks)

        return company, df_monthly, df_partners, df_tasks

    except Exception as e:
        st.error(f"데이터 로딩 오류: {e}")
        return None, None, None, None

st.markdown('<div class="main-header">💰 물류비 통합 관리 시스템</div>', unsafe_allow_html=True)

# 사이드바 - 파일 선택
st.sidebar.header("📁 파일 선택")

file_option = st.sidebar.radio(
    "데이터 소스:",
    ["기본 파일", "파일 업로드", "경로 입력"],
    label_visibility="collapsed"
)

excel_file = None

if file_option == "기본 파일":
    excel_file = "물류비_관리대장.xlsx"
    st.sidebar.success("✅ 기본 파일 사용 중")

elif file_option == "파일 업로드":
    uploaded_file = st.sidebar.file_uploader(
        "엑셀 파일 업로드",
        type=['xlsx'],
        help="물류비_관리대장.xlsx 형식의 파일"
    )
    if uploaded_file:
        # 임시 파일로 저장
        import tempfile
        import os
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            tmp_file.write(uploaded_file.getvalue())
            excel_file = tmp_file.name
        st.sidebar.success(f"✅ {uploaded_file.name}")
    else:
        st.sidebar.info("파일을 업로드하세요")

elif file_option == "경로 입력":
    custom_path = st.sidebar.text_input(
        "파일 경로:",
        value="물류비_관리대장.xlsx",
        help="예: C:/data/물류비.xlsx"
    )
    if custom_path:
        import os
        if os.path.exists(custom_path):
            excel_file = custom_path
            st.sidebar.success(f"✅ 파일 찾음")
        else:
            st.sidebar.error("❌ 파일 없음")

# 데이터 로드
if excel_file:
    company, df_monthly, df_partners, df_tasks = load_logistics_data(excel_file)
else:
    company, df_monthly, df_partners, df_tasks = None, None, None, None

if company is None:
    st.error("엑셀 파일을 찾을 수 없습니다!")
    st.stop()

st.sidebar.markdown("---")
st.sidebar.header("🏢 회사 정보")
st.sidebar.info(f"""**{company['회사명']}**
업종: {company['업종']}
연매출: {company['연매출']:,.0f}억원
목표 물류비율: {company['목표물류비율']:.1f}%""")

st.sidebar.markdown("---")
st.sidebar.header("📊 데이터 현황")
st.sidebar.metric("월별 데이터", f"{len(df_monthly)}개월")
st.sidebar.metric("협력사", f"{len(df_partners)}개")
st.sidebar.metric("개선 과제", f"{len(df_tasks)}건")

st.sidebar.markdown("---")
if st.sidebar.button("🔄 데이터 새로고침", use_container_width=True):
    st.cache_data.clear()
    st.rerun()

tab1, tab2, tab3, tab4, tab5 = st.tabs(["📊 대시보드", "📈 월별 분석", "🤝 협력사 관리", "💡 개선 과제", "📋 리포트"])

with tab1:
    st.header("📊 종합 대시보드")

    latest = df_monthly.iloc[-1]

    col1, col2, col3, col4 = st.columns(4)

    with col1:
        delta = latest['총물류비'] - df_monthly.iloc[-2]['총물류비'] if len(df_monthly) > 1 else 0
        st.metric("최근 월 물류비", f"{latest['총물류비']:,.0f}만원", f"{delta:+,.0f}만원")

    with col2:
        delta_ratio = latest['물류비율'] - company['목표물류비율']
        st.metric("물류비율", f"{latest['물류비율']:.1f}%", f"{delta_ratio:+.1f}%p", delta_color="inverse")

    with col3:
        avg_3m = df_monthly.tail(3)['물류비율'].mean()
        st.metric("3개월 평균", f"{avg_3m:.1f}%")

    with col4:
        benchmark = 16.0
        diff = latest['물류비율'] - benchmark
        st.metric("업종 평균 대비", f"{diff:+.1f}%p", delta_color="inverse" if diff > 0 else "normal")

    if latest['물류비율'] > company['목표물류비율'] * 1.2:
        st.markdown(f'<div class="alert-danger"><strong>🚨 긴급 조치 필요</strong><br>물류비율 {latest["물류비율"]:.1f}%가 목표 {company["목표물류비율"]:.1f}%를 크게 초과</div>', unsafe_allow_html=True)
    elif latest['물류비율'] > company['목표물류비율']:
        st.markdown(f'<div class="alert-warning"><strong>⚠️ 주의 필요</strong><br>목표 대비 {latest["물류비율"] - company["목표물류비율"]:.1f}%p 초과</div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="alert-success"><strong>✅ 양호</strong><br>물류비율이 목표 범위 내</div>', unsafe_allow_html=True)

    st.markdown("---")

    col1, col2 = st.columns(2)

    with col1:
        st.subheader("📈 월별 물류비 추이")
        fig1 = go.Figure()
        fig1.add_trace(go.Scatter(x=df_monthly['년월'], y=df_monthly['총물류비'], mode='lines+markers', line=dict(color='#1F4E78', width=3), marker=dict(size=8)))
        fig1.update_layout(height=400, yaxis_title="물류비 (만원)")
        st.plotly_chart(fig1, use_container_width=True)

    with col2:
        st.subheader("💼 최근 물류비 구성")
        items = ['국제운송비', '창고비', '국내배송비', '포장비', '통관비', '기타']
        values = [latest[i] for i in items]
        colors = ['#DC2626', '#F59E0B', '#3B82F6', '#8B5CF6', '#10B981', '#6366F1']
        fig2 = go.Figure(data=[go.Pie(labels=items, values=values, hole=.4, marker=dict(colors=colors))])
        fig2.update_layout(height=400)
        st.plotly_chart(fig2, use_container_width=True)

with tab2:
    st.header("📈 월별 상세 분석")

    st.subheader("📋 월별 데이터")
    st.dataframe(df_monthly.style.format({'국제운송비': '{:,.0f}', '창고비': '{:,.0f}', '국내배송비': '{:,.0f}', '포장비': '{:,.0f}', '통관비': '{:,.0f}', '기타': '{:,.0f}', '총물류비': '{:,.0f}', '물류비율': '{:.2f}%'}), use_container_width=True, hide_index=True, height=300)

    st.markdown("---")

    col1, col2 = st.columns(2)

    with col1:
        st.subheader("📊 항목별 통계")
        cost_items = ['국제운송비', '창고비', '국내배송비', '포장비', '통관비', '기타']
        stats = df_monthly[cost_items].describe().T[['mean', 'min', 'max']]
        stats.columns = ['평균', '최소', '최대']
        st.dataframe(stats.style.format('{:,.0f}'), use_container_width=True)

    with col2:
        st.subheader("📈 전월 대비 증감")
        if len(df_monthly) >= 2:
            latest = df_monthly.iloc[-1]
            prev = df_monthly.iloc[-2]
            changes = []
            for item in cost_items:
                delta = latest[item] - prev[item]
                pct = (delta / prev[item] * 100) if prev[item] > 0 else 0
                changes.append({'항목': item, '증감액': f"{delta:+,.0f}", '증감률': f"{pct:+.1f}%"})
            st.dataframe(pd.DataFrame(changes), use_container_width=True, hide_index=True)

    st.markdown("---")
    st.subheader("📊 항목별 비용 추이")
    fig = go.Figure()
    colors = ['#DC2626', '#F59E0B', '#3B82F6', '#8B5CF6', '#10B981', '#6366F1']
    for item, color in zip(cost_items, colors):
        fig.add_trace(go.Bar(name=item, x=df_monthly['년월'], y=df_monthly[item], marker_color=color))
    fig.update_layout(barmode='stack', height=450, yaxis_title="비용 (만원)")
    st.plotly_chart(fig, use_container_width=True)

with tab3:
    st.header("🤝 협력사 관리")
    st.success(f"총 {len(df_partners)}개 협력사")

    for _, row in df_partners.iterrows():
        with st.expander(f"**{row['구분']}** - {row['협력사명']}"):
            col1, col2, col3 = st.columns(3)
            with col1:
                st.write(f"**월평균**: {row['월평균비용']:,.0f}만원")
                st.write(f"**담당자**: {row['담당자']}")
            with col2:
                st.write(f"**계약 시작**: {row['계약시작일']}")
                st.write(f"**계약 종료**: {row['계약종료일']}")
            with col3:
                st.write(f"**연락처**: {row['연락처']}")

with tab4:
    st.header("💡 개선 과제 관리")

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("전체 과제", f"{len(df_tasks)}건")
    with col2:
        target = df_tasks['목표절감액'].sum()
        st.metric("목표 절감액", f"{target:,.0f}만원/년")
    with col3:
        actual = df_tasks['실제절감액'].sum()
        st.metric("실제 절감액", f"{actual:,.0f}만원/년")
    with col4:
        achievement = (actual / target * 100) if target > 0 else 0
        st.metric("달성률", f"{achievement:.1f}%")

    st.markdown("---")

    for _, row in df_tasks.iterrows():
        emoji = {"진행중": "🔄", "완료": "✅", "예정": "📅"}.get(row['상태'], "⚪")
        with st.expander(f"{emoji} {row['과제명']} ({row['상태']})"):
            col1, col2 = st.columns(2)
            with col1:
                st.write(f"**카테고리**: {row['카테고리']}")
                st.write(f"**담당자**: {row['담당자']}")
                st.write(f"**기간**: {row['시작일']} ~ {row['완료예정일']}")
            with col2:
                st.write(f"**목표**: {row['목표절감액']:,.0f}만원/년")
                st.write(f"**실제**: {row['실제절감액']:,.0f}만원/년")
                if row['목표절감액'] > 0:
                    task_ach = (row['실제절감액'] / row['목표절감액'] * 100)
                    st.progress(min(task_ach / 100, 1.0))
                    st.write(f"달성률: {task_ach:.1f}%")

with tab5:
    st.header("📋 종합 리포트")

    latest = df_monthly.iloc[-1]

    report = f"""## 물류비 관리 종합 리포트

**생성일시**: {datetime.now().strftime('%Y-%m-%d %H:%M')}
**회사명**: {company['회사명']}
**업종**: {company['업종']}

---

### 📊 주요 지표
- **최근 월({latest['년월']}) 물류비**: {latest['총물류비']:,.0f}만원
- **물류비율**: {latest['물류비율']:.2f}%
- **목표 대비**: {latest['물류비율'] - company['목표물류비율']:+.2f}%p

### 💡 개선 과제 현황
- **전체 과제**: {len(df_tasks)}건
- **목표 절감액**: {df_tasks['목표절감액'].sum():,.0f}만원/년
- **실제 절감액**: {df_tasks['실제절감액'].sum():,.0f}만원/년
- **달성률**: {(df_tasks['실제절감액'].sum() / df_tasks['목표절감액'].sum() * 100):.1f}%

### 🤝 협력사 현황
- **등록 협력사**: {len(df_partners)}개
- **월평균 총 비용**: {df_partners['월평균비용'].sum():,.0f}만원
"""

    st.markdown(report)
    st.markdown("---")

    col1, col2, col3 = st.columns(3)
    with col1:
        st.download_button("📄 리포트 다운로드", report, f"물류비리포트_{datetime.now().strftime('%Y%m%d')}.txt", use_container_width=True)
    with col2:
        csv = df_monthly.to_csv(index=False, encoding='utf-8-sig')
        st.download_button("📊 월별 데이터", csv, f"월별물류비_{datetime.now().strftime('%Y%m%d')}.csv", use_container_width=True)
    with col3:
        csv_p = df_partners.to_csv(index=False, encoding='utf-8-sig')
        st.download_button("🤝 협력사 데이터", csv_p, f"협력사_{datetime.now().strftime('%Y%m%d')}.csv", use_container_width=True)
